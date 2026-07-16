"""
Extract priors from Indian flight dataset (offline tool).

Run this once to generate config/indian_priors.json.
The JSON file is then shipped with the app for cold-start usage.

Usage:
    python tools/extract_indian_priors.py <path_to_Data_Train.xlsx>
    
Output:
    config/indian_priors.json
"""
import json
import logging
import sys
from pathlib import Path
from typing import Dict, List

# Add project root
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from flight_monitor.core.predictor.indian_prior import IndianPriorExtractor

logging.basicConfig(level=logging.INFO, format="%(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


def load_records(xlsx_path: str) -> List[Dict]:
    """Load Indian flight records from Excel file."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        airline, date_str, src, dst, route, dep_time, arr_time, duration, stops, info, price = row
        if price and price > 0:
            records.append({
                "airline": airline,
                "date": str(date_str) if date_str else "",
                "source": src,
                "destination": dst,
                "route": route or "",
                "dep_time": dep_time or "",
                "arrival_time": arr_time or "",
                "duration": duration or "",
                "stops": stops,
                "additional_info": info or "",
                "price": float(price),
            })

    return records


def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_indian_priors.py <path_to_Data_Train.xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not Path(xlsx_path).exists():
        print(f"ERROR: File not found: {xlsx_path}")
        sys.exit(1)

    # Load records
    logger.info(f"Loading records from {xlsx_path}...")
    records = load_records(xlsx_path)
    logger.info(f"Loaded {len(records)} price records")

    # Extract priors
    extractor = IndianPriorExtractor()
    priors = extractor.extract_from_records(records)

    # Save
    output_dir = Path(__file__).resolve().parent.parent / "flight_monitor" / "config"
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / "indian_priors.json"

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(priors, f, ensure_ascii=False, indent=2, default=str)

    logger.info(f"Priors saved to {output_path}")
    print(f"Done. Extract complete: {output_path}")
    print(f"Records: {priors['n_price_records']}")
    print(f"Stop discount: {priors['stop_discount']['average_stop_discount']}")
    print(f"Tiers: {list(priors['airline_classification']['tier_stats'].keys())}")


if __name__ == "__main__":
    from typing import List, Dict
    main()
